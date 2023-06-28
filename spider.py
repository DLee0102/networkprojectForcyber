from openpyxl import Workbook
from openpyxl import load_workbook

from bs4 import BeautifulSoup      #网页解析，获取数据
import re       #正则表达式，进行文字匹配
import urllib.request, urllib.error     #制定URL，获取网页数据
import sqlite3  #进行SQLite数据库操作
import os

# 注意：正则匹配要以接收到的文本为依据，不能以浏览器开发工具中看到的文本为依据，因为有时候两者内容不一样
# 要常加?和re.S来进行正则匹配（多行匹配）

dbpath = './movie_total.db'

# 影片图片
findImgsrc = re.compile(r'<img.*src="(.*?)" title="点击看更多海报"',re.S)   #问好表示非贪婪模式，即匹配尽可能少的内容

# 影片片名
findTitle = re.compile(r'<span property="v:itemreviewed">(.*?)</span>')

# 剧情简介
findIntro_simple = re.compile(r'<span class="" property="v:summary">(.*?)</span>', re.S)
findIntro_hidden = re.compile(r'<span class="all hidden">(.*?)</span>', re.S)

# 找到导演
finddirector = re.compile(r'<a href=".*" rel="v:directedBy">(.*?)</a>')
# 找到编剧
findeditors = re.compile(r'<span class="attrs">(.*?)</span>', re.S)
findeditor = re.compile(r'<a href=".*">(.*)</a>')
# 找到演员 
findactors = re.compile(r'<span class="attrs">(.*?)</span>', re.S)
findactor = re.compile(r'<a href=".*" rel="v:starring">(.*?)</a>', re.S)
# 找到类型
findclasses = re.compile(r'<span property="v:genre">(.*?)</span>', re.S)
# 找到国家地区
findlocation = re.compile(r'<span class="pl">制片国家/地区:</span>(.*?)<br/>', re.S)
# 找到语言
findlanguage = re.compile(r'<span class="pl">语言:</span>(.*?)<br/>')
# 找到上映时间（有两种格式的上映时间）
findtimes = re.compile(r'<span class="pl">上映日期:</span>(.*?)<br/>', re.S)
findtime = re.compile(r'<span content=".*" property="v:initialReleaseDate">(\d{4}-\d{1,2}-\d{1,2}).*</span>', re.S)
findtime_loc = re.compile(r'<span content=".*" property="v:initialReleaseDate">\d{4}-\d{1,2}-\d{1,2}(.*?)</span>', re.S)
findtime_ = re.compile(r'<span content=".*" property="v:initialReleaseDate">(\d*).*</span>', re.S)
findtime_loc_ = re.compile(r'<span content=".*" property="v:initialReleaseDate">\d*(.*?)</span>', re.S)
# 找到片长
findlength = re.compile(r'<span content=".*" property="v:runtime">(\d*).*</span>')
# 找到又名
findanothername = re.compile(r'<span class="pl">又名:</span>(.*?)<br/>')
# 找到IMDb
findIMDb = re.compile(r'<span class="pl">IMDb:</span>(.*?)<br/>')
# 找到评分
findscores = re.compile(r'<strong class="ll rating_num" property="v:average">(.*?)</strong>')
# 找到评价人数
findcommentcounts = re.compile(r'<span property="v:votes">(.*?)</span>人评价')
# 找到短评和影评人数
findreviewcounts = re.compile(r'<a href=".*">全部 (.*?) 条</a>')

# 找到短评信息
# 找到短评昵称
findcommentator = re.compile(r'<a class="" href=".*">(.*?)</a>')
# 找到短评时间
commenttimepattern = r"\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}"
# 找到短评内容
findcommentcontent = re.compile(r'<span class="short">(.*?)</span>', re.S)
# 找到认为短评有用人数
findcommentuseful = re.compile(r'<span class="votes vote-count">(.*?)</span>')

# 找到影评信息
# 找到影评者
findreviewer = re.compile(r'<a class="name" href=".*">(.*?)</a>')
# 找到影评时间
findreviewtime = re.compile(r'<span class="main-meta" content=".*">(.*?)</span>')
# 找到影评内容
findreviewcontent = re.compile(r'<div class="short-content">(.*?)</div>', re.S)
findreviewrealcon = re.compile(r'\n\n(.*?)\n\n', re.S)
# 找到认为有用人数
findreviewuseful = re.compile(r'<span id="r-useful_count-.*">\s*(\d+)\s*</span>')
# 找到认为无用人数
findreviewuseless = re.compile(r'<span id="r-useless_count-.*>(.*?)</span>', re.S)
# 找到回应人数
findreviewresponse = re.compile(r'<a class="reply" href=".*">(.*?)回应</a>')


def readfromExcel(wbname='doubanMovielinks.xlsx', sheetname='Sheet',
                    start_row=4004, finish_row=4501):
    link_list = []
    wb = load_workbook(wbname)
    ws = wb.get_sheet_by_name(sheetname)
    for row in ws.iter_rows(min_row=start_row, max_col=1, max_row=finish_row):
        for cell in row:
            link_list.append(cell.value)
    
    return link_list

def askURL(url):
    head = {    #模拟浏览器信息，向豆瓣服务器发送消息
        "User-Agent":"Mozilla / 5.0(Windows NT 10.0;Win64;x64) AppleWebKit / 537.36(KHTML, likeGecko) Chrome / 97.0.4692.99 Safari / 537.36 Edg / 97.0.1072.76"
    }
    #用户代理，表示告诉服务器我们的机器类型（本质上告诉浏览器我们可以接受什么水平的内容）

    #发送消息
    request = urllib.request.Request(url = url, headers = head)
    html = ""
    try:
        response = urllib.request.urlopen(request)  #接收返回的信息
        html = response.read().decode("utf-8")  #读取信息
        # print(html)
    except urllib.error.URLError as e:
        if hasattr(e,"code"):       #该函数用于判断是否含指定字符串
            print(e.code)       #用上面的函数打印指定字符串
        if hasattr(e,"reason"):
            print(e.reason)
    return html
    
# 获取数据
def getData(link_list):
    datalist = []
    for i, link_ in enumerate(link_list):
        try:
            html = askURL(link_)
            data = []
            
            soup = BeautifulSoup(html, "html.parser")
            
            # 找到电影标题
            for title in soup.find_all('span', property="v:itemreviewed"):
                title = str(title)
                
                titlename = re.findall(findTitle, title)[0]
                if len(titlename.split(" ", 1)) == 1:
                    cname = titlename.split(" ", 1)[0]
                    fname = ''
                else:
                    cname = titlename.split(" ", 1)[0]
                    fname = titlename.split(" ", 1)[1]
                data.append(cname)
                data.append(fname)
            
            for basic in soup.find_all('div', 'subjectwrap clearfix'):
                basic = str(basic)
                
                # 找海报
                imgsrc = re.findall(findImgsrc, basic)[0]
                data.append(imgsrc)
                
                # 找导演
                director = re.findall(finddirector, basic)
                data.append(director[0])
                
                # 找编剧（将编剧信息保存为列表）
                editors = re.findall(findeditors, basic)
                editor = []
                editors = editors[1].split(' / ')
                # print(editors)
                for e in editors:
                    editor.append(re.findall(findeditor, e)[0])
                # print(editor)
                data.append(editor)
                
                # 找演员（将演员信息保存为列表）
                actors = re.findall(findactors, basic)
                actor = []
                # print(actors[2])
                for a in actors[2].split(' / '):
                    # print(re.findall(findactor, a))
                    actor.append(re.findall(findactor, a)[0])
                data.append(actor)
                
                # 找类型
                classes = re.findall(findclasses, basic)
                class_string = ""
                for class_name in classes:
                    class_string += class_name + " / "
                data.append(class_string)
                
                # 找到国家地区
                # print(basic)
                location = re.findall(findlocation, basic)
                data.append(location[0])
                
                # 找到语言
                language = re.findall(findlanguage, basic)
                data.append(language[0])
                
                # 找到上映时间（处理不同格式的上映时间）
                time = []
                for t in re.findall(findtimes, basic)[0].split(" / "):
                    # print(re.findall(findtimes, basic))
                    if re.findall(findtime, t)==[]:
                        time.append([re.findall(findtime_, t)[0], re.findall(findtime_loc_, t)[0]])  # 格式为1984（）
                    else:
                        time.append([re.findall(findtime, t)[0], re.findall(findtime_loc, t)[0]])  # 格式为1984-8-2（）
                data.append(time)
                
                # 找到片长
                length = re.findall(findlength, basic)
                data.append(length[0])
                
                # 找到又名
                anothername = re.findall(findanothername, basic)
                # 又名可能为空
                if anothername == []:
                    data.append("  ")
                else:
                    data.append(anothername[0])
                    
                # 找到IMDb
                IMDbinfo = re.findall(findIMDb, basic)
                IMDbinfo[0] = IMDbinfo[0].strip()
                data.append(IMDbinfo[0])
                
                # 找到评分和评价人数
                score = re.findall(findscores, basic)
                commentcounts = re.findall(findcommentcounts, basic)
                data.append(score[0])
                data.append(commentcounts[0])
                
            # 找到剧情简介
            for relatedinfo in soup.find_all('div', "related-info"):
                relatedinfo = str(relatedinfo)
                intro = re.findall(findIntro_hidden, relatedinfo)
                if intro == []:
                    intro = re.findall(findIntro_simple, relatedinfo)
                intro_single = intro[0]
                intro_single = re.sub('\s+', " ", intro_single).replace(r"<br/>", "\n", 10)
                data.append(intro_single)
                
            for review_count in soup.find_all('span', 'pl'):
                review_count = str(review_count)
                review_c = re.findall(findreviewcounts, review_count)
                if review_c != []:
                    data.append(review_c[0])
                
            # 找到短评信息
            comments_list = []
            # 删除最后一个元素
            for commentinfo in soup.find_all('div', 'comment-item')[:-1]:
                comments = []
                commentinfo = str(commentinfo)
                # print(commentinfo)
                
                # 找到评论者
                commentator = re.findall(findcommentator, commentinfo)
                comments.append(commentator[0])
                
                # 找到评论时间
                commenttime = re.findall(commenttimepattern, commentinfo)
                comments.append(commenttime[0])
                
                # 找到评论内容
                commentcontent = re.findall(findcommentcontent, commentinfo)
                comments.append(commentcontent[0])
                
                # 找到认为短评有用人数
                commentuseful = re.findall(findcommentuseful, commentinfo)
                comments.append(commentuseful[0])
                
                comments_list.append(comments)
            
            # 找到影评信息
            review_list = []
            for reviewinfo in soup.find_all('div', 'main review-item'):
                reviews = []
                reviewinfo = str(reviewinfo)
                # print(reviewinfo)
                
                # 找到影评者
                reviewer = re.findall(findreviewer, reviewinfo)
                reviews.append(reviewer[0])
                
                # 找到影评时间
                reviewtime = re.findall(findreviewtime, reviewinfo)
                reviews.append(reviewtime[0])
                
                # 找到影评内容
                reviewcontent = re.findall(findreviewcontent, reviewinfo)
                reviewcontent = re.findall(findreviewrealcon, reviewcontent[0])
                reviewcontent = re.sub('\s+', " ", reviewcontent[0])
                reviews.append(reviewcontent)
                
                # 找到认为有用人数
                reviewuseful = re.findall(findreviewuseful, reviewinfo)
                if reviewuseful == []:
                    reviewuseful.append('0')
                reviewuseful = re.sub('\s+', "", reviewuseful[0]).replace(r"<br/>", "\n", 10)
                reviews.append(reviewuseful)
                
                # 找到认为无用人数
                reviewuseless = re.findall(findreviewuseless, reviewinfo)
                if reviewuseless == []:
                    reviewuseless.append('0')
                reviewuseless = re.sub('\s+', "", reviewuseless[0]).replace(r"<br/>", "\n", 10)
                reviews.append(reviewuseless)
                
                # 找到回应人数
                reviewresponse = re.findall(findreviewresponse, reviewinfo)
                reviews.append(reviewresponse[0])
                
                review_list.append(reviews)
            
            print("爬取第" + str(i+1) + "条链接：" + link_ + " 成功")
            datalist.append([data, comments_list, review_list])
            # print(data)
            # print(comments_list)
            # print(review_list)
        except IndexError as e:
            print("警告：缺少主键参数", e, "跳过该条电影")
        
    print("---------------------------------")
    print("------------ 全部爬取完毕")
    return datalist

def saveData2DB(datalist,dbpath):
    print("------------------------------------")
    print("------- 正在将爬取内容添加至数据库...")
    conn = sqlite3.connect(dbpath)
    cur = conn.cursor()
    for item in datalist:
        try:
            basic_list = []
            for index, meta in enumerate(item[0]):
                if index == 4:
                    for oneeditor in meta:
                        sql_editorsIMDb = '"' + item[0][12] + '"'
                        sql_editorseditor_name = '"' + oneeditor + '"'
                        sql_editors_value = [sql_editorsIMDb, sql_editorseditor_name]
                        sql_editors = '''
                        insert into editors (
                            IMDb,
                            editor_name
                        )values (%s)
                        '''%",".join(sql_editors_value)
                        cur.execute(sql_editors)
                else:
                    if index == 5:
                        for oneactor in meta:
                            sql_actorsIMDb = '"' + item[0][12] + '"'
                            sql_actorsactor_name = '"' + oneactor + '"'
                            sql_actors_value = [sql_actorsIMDb, sql_actorsactor_name]
                            sql_actors = '''
                            insert into actors (
                                IMDb,
                                actor_name
                            )values (%s)
                            '''%",".join(sql_actors_value)
                            cur.execute(sql_actors)
                    else:
                        if index == 9:
                            basic_list.append('"' + meta[0][0] + '"')
                        else:
                            basic_list.append('"' + meta.replace('"', '""').replace("'", "''") + '"')
                    
            sql_basic = '''
            insert into basic (
                cname,
                fname,
                pic_link,
                director,
                classes,
                location,
                language,
                uptime,
                length,
                other_name,
                IMDb,
                score,
                rated,
                instruction,
                comments_count,
                reviews_count
            )values (%s)
            '''%",".join(basic_list)
            cur.execute(sql_basic)
            
            # 插入短评表
            for index, meta in enumerate(item[1]):
                onecomment = []
                onecomment.append('"' + item[0][12] + '"')
                # 转义双引号和单引号避免sql错误
                onecomment.append('"' + meta[0].replace('"', '""').replace("'", "''") + '"')
                onecomment.append('"' + meta[1].replace('"', '""').replace("'", "''") + '"')
                onecomment.append('"' + meta[2].replace('"', '""').replace("'", "''") + '"')
                onecomment.append('"' + meta[3].replace('"', '""').replace("'", "''") + '"')
                
                sql_comments = '''
                insert into comments (
                    IMDb,
                    nickname,
                    commenttime,
                    content,
                    count_useful
                )values (%s)
                '''%",".join(onecomment)
                cur.execute(sql_comments)
                
            # 插入影评表
            for index, meta in enumerate(item[2]):
                onecomment = []
                onecomment.append('"' + item[0][12] + '"')
                # 转义双引号和单引号避免sql错误
                onecomment.append('"' + meta[0].replace('"', '""').replace("'", "''") + '"')
                onecomment.append('"' + meta[1].replace('"', '""').replace("'", "''") + '"')
                onecomment.append('"' + meta[2].replace('"', '""').replace("'", "''") + '"')
                onecomment.append('"' + meta[3].replace('"', '""').replace("'", "''") + '"')
                onecomment.append('"' + meta[4].replace('"', '""').replace("'", "''") + '"')
                onecomment.append('"' + meta[5].replace('"', '""').replace("'", "''") + '"')
                
                sql_comments = '''
                insert into reviews (
                    IMDb,
                    nickname,
                    commenttime,
                    content,
                    count_useful,
                    count_useless,
                    count_response
                )values (%s)
                '''%",".join(onecomment)
                cur.execute(sql_comments)
        except sqlite3.OperationalError as e0:
            print("警告：插入数据库错误", e0, "跳过该条电影")
        except sqlite3.IntegrityError as e1:
            print("警告：数据库完整性错误", e1, "跳过该条电影")
        except ValueError as e2:
            print(e2)
    
    conn.commit()
    cur.close()
    conn.close()
    print("----------- 添加成功！")
                

def main():
    link_list = readfromExcel()
    datalist = getData(link_list)
    saveData2DB(datalist, dbpath)



if __name__ == '__main__':
    main()